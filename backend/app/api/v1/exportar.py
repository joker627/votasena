# ==============================================================================
# MÓDULO: Generación y Exportación de Reportes (Excel & PDF)
# ==============================================================================

import io
import pymysql
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from app.core.database import get_db_connection
from app.api.v1.auth import get_current_admin

router = APIRouter(
    prefix="/exportar",
    tags=["exportar"]
)

# --- Consultas de Datos Consolidados ---
def obtener_resultados_completos(jornada: str = None):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Error de base de datos")
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            if jornada and jornada != 'Todas':
                cursor.execute("""
                    SELECT c.nombre, c.numero_tarjeton, c.jornada, COUNT(v.id) as total_votos
                    FROM candidatos c
                    LEFT JOIN votos v ON c.id = v.candidato_id
                    WHERE c.jornada = %s
                    GROUP BY c.id
                    ORDER BY c.jornada, total_votos DESC
                """, (jornada,))
            else:
                cursor.execute("""
                    SELECT c.nombre, c.numero_tarjeton, c.jornada, COUNT(v.id) as total_votos
                    FROM candidatos c
                    LEFT JOIN votos v ON c.id = v.candidato_id
                    GROUP BY c.id
                    ORDER BY c.jornada, total_votos DESC
                """)
            return cursor.fetchall()
    finally:
        conn.close()

# --- Endpoints de Exportación ---
@router.get("/excel")
def exportar_excel(jornada: str = None, admin: dict = Depends(get_current_admin)):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    resultados = obtener_resultados_completos(jornada)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Resultados {jornada}" if jornada and jornada != 'Todas' else "Resultados Oficiales"
    
    # Encabezados
    ws.append(["Candidato", "Nº Tarjetón", "Jornada", "Votos Obtenidos"])
    
    # Filas de datos
    for r in resultados:
        ws.append([r['nombre'], r['numero_tarjeton'], r['jornada'], r['total_votos']])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Resultados_VotaSena.xlsx"}
    )

@router.get("/pdf")
def exportar_pdf(jornada: str = None, admin: dict = Depends(get_current_admin)):
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 no instalado")

    resultados = obtener_resultados_completos(jornada)
    
    pdf = FPDF()
    pdf.add_page()
    
    # Paleta de color institucional
    verde_sena = (57, 169, 0)
    gris_oscuro = (50, 50, 50)
    
    # Franja superior
    pdf.set_fill_color(*verde_sena)
    pdf.rect(0, 0, 210, 15, 'F')
    
    # Títulos
    pdf.set_y(25)
    pdf.set_font("Arial", style='B', size=18)
    pdf.set_text_color(*verde_sena)
    pdf.cell(0, 10, "SENA - SISTEMA INTEGRADO DE ELECCIONES", align='C', ln=True)
    
    pdf.set_font("Arial", size=12)
    pdf.set_text_color(100, 100, 100)
    titulo_jornada = f"Jornada: {jornada}" if jornada and jornada != 'Todas' else "Consolidado General"
    pdf.cell(0, 10, f"Reporte Oficial de Resultados | {titulo_jornada}", align='C', ln=True)
    
    pdf.ln(10)
    
    # Cabecera de tabla
    pdf.set_font("Arial", style='B', size=11)
    pdf.set_fill_color(240, 240, 240)
    pdf.set_text_color(*gris_oscuro)
    pdf.set_draw_color(200, 200, 200)
    
    w = [80, 25, 40, 40]
    headers = ["Candidato", "No.", "Jornada", "Total Votos"]
    for i in range(len(headers)):
        pdf.cell(w[i], 12, headers[i], border=1, fill=True, align='C')
    pdf.ln()
    
    # Filas de datos
    pdf.set_font("Arial", size=10)
    fill = False
    pdf.set_fill_color(250, 250, 250)
    
    for r in resultados:
        nombre = str(r['nombre']).encode('latin-1', 'replace').decode('latin-1')
        jornada_txt = str(r['jornada']).encode('latin-1', 'replace').decode('latin-1')
        votos = str(r['total_votos'])
        num = str(r['numero_tarjeton'])
        
        pdf.cell(w[0], 12, f"  {nombre}", border=1, fill=fill, align='L')
        pdf.cell(w[1], 12, num, border=1, fill=fill, align='C')
        pdf.cell(w[2], 12, jornada_txt, border=1, fill=fill, align='C')
        
        pdf.set_font("Arial", style='B', size=11)
        pdf.set_text_color(*verde_sena)
        pdf.cell(w[3], 12, votos, border=1, fill=fill, align='C')
        
        pdf.set_font("Arial", size=10)
        pdf.set_text_color(*gris_oscuro)
        fill = not fill
        pdf.ln()

    pdf.ln(15)
    pdf.set_font("Arial", style='I', size=9)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, "Generado automáticamente por el Sistema de Votación VotaSena.", align='C')
        
    output = io.BytesIO(bytes(pdf.output()))
    
    return StreamingResponse(
        output, 
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=Resultados_VotaSena.pdf"}
    )
